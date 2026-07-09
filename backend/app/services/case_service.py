"""测试用例（覆盖率分母）业务逻辑：查询/新增/Excel 导入导出/模板。"""
import io
from contextlib import closing
from datetime import datetime

from openpyxl import Workbook, load_workbook

from .. import db

# 导入模板列（与原型导入弹窗说明一致）；自动化状态由脚本映射判定，不在模板内
TEMPLATE_HEADERS = ["用例编号", "所属App", "功能模块", "优先级"]

# 用例的「关联脚本」：一条用例可能被多个脚本覆盖，展示最新一个（与原型单列展示一致）
_SCRIPT_SUB = ("(SELECT s.name FROM script_case sc JOIN scripts s ON s.id = sc.script_id "
               " WHERE sc.case_id = c.id ORDER BY s.id DESC LIMIT 1)")


def list_cases(app: str = "", priority: str = "", status: str = "") -> dict:
    """按 App / 优先级 / 自动化状态 筛选。返回 {total, rows}。

    自动化状态不是存储字段：有脚本映射 = 已自动化，否则 = 待自动化。
    """
    sql = (f"SELECT c.id, c.app, c.module, c.priority, c.title, {_SCRIPT_SUB} AS script"
           " FROM cases c")
    cond, args = [], []
    if app and app != "全部 App":
        cond.append("c.app = ?"); args.append(app)
    if priority and priority != "全部优先级":
        cond.append("c.priority = ?"); args.append(priority)
    if status == "已自动化":
        cond.append(f"{_SCRIPT_SUB} IS NOT NULL")
    elif status == "待自动化":
        cond.append(f"{_SCRIPT_SUB} IS NULL")
    if cond:
        sql += " WHERE " + " AND ".join(cond)
    # 新增/导入的用例（created_at 更新）排前，同批内保持录入顺序
    sql += " ORDER BY c.created_at DESC, c.rowid ASC"
    with closing(db.get_conn()) as conn:
        rows = [dict(r) for r in conn.execute(sql, args)]
        total = conn.execute("SELECT COUNT(*) FROM cases").fetchone()[0]
    return {"total": total, "rows": rows}


def create_case(cid: str, app: str, module: str, priority: str):
    """新增功能用例；编号重复时抛 ValueError（前端 toast『用例编号已存在』）。"""
    with closing(db.get_conn()) as conn:
        if conn.execute("SELECT 1 FROM cases WHERE id = ?", (cid,)).fetchone():
            raise ValueError("用例编号已存在")
        conn.execute("INSERT INTO cases(id,app,module,priority,created_at) VALUES(?,?,?,?,?)",
                     (cid, app, module, priority, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit()


def import_excel(content: bytes, dry_run: bool = False) -> dict:
    """Excel 导入：以「用例编号」为主键 upsert（存在则更新、不存在则新增）。

    dry_run=True 只解析统计不落库，用于导入弹窗的预览行为（与原型一致：
    选完文件先显示「已解析 N 条 · 新增 X / 更新 Y」，点确定才真正导入）。
    """
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header or [str(h or "").strip() for h in header[:4]] != TEMPLATE_HEADERS:
        raise ValueError("表头不符合模板，请先下载导入模板")

    parsed, errors = [], []
    for i, row in enumerate(rows_iter, start=2):        # Excel 行号从 2 起（1 为表头）
        cid, app, module, pri = [str(v).strip() if v is not None else "" for v in (row or ("",) * 4)[:4]]
        if not any([cid, app, module, pri]):
            continue                                     # 跳过空行
        if not cid or not app or not module:
            errors.append(f"第 {i} 行缺少必填列")
            continue
        parsed.append((cid, app, module, pri or "P1"))

    added = updated = 0
    with closing(db.get_conn()) as conn:
        exist = {r[0] for r in conn.execute("SELECT id FROM cases")}
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for cid, app, module, pri in parsed:
            if cid in exist:
                updated += 1
                if not dry_run:
                    conn.execute("UPDATE cases SET app=?, module=?, priority=? WHERE id=?",
                                 (app, module, pri, cid))
            else:
                added += 1
                exist.add(cid)                           # 同文件内重复编号按更新计
                if not dry_run:
                    conn.execute("INSERT INTO cases(id,app,module,priority,created_at) VALUES(?,?,?,?,?)",
                                 (cid, app, module, pri, ts))
        if not dry_run:
            conn.commit()
    return {"parsed": len(parsed), "added": added, "updated": updated, "errors": errors}


def _build_workbook(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "功能用例"
    ws.append(TEMPLATE_HEADERS + ["自动化状态", "关联脚本"])
    for r in rows:
        ws.append([r["id"], r["app"], r["module"], r["priority"],
                   "已自动化" if r["script"] else "待自动化", r["script"] or ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_excel(app: str = "", priority: str = "", status: str = "") -> bytes:
    """导出当前筛选结果为 .xlsx（评审/备份用）。"""
    return _build_workbook(list_cases(app, priority, status)["rows"])


def template_excel() -> bytes:
    """导入模板：表头 + 一行示例。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    ws.append(TEMPLATE_HEADERS)
    ws.append(["KW-PLAY-010", "酷我音乐", "播放控制", "P0"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

"""官方用例导入 + 框架适配器（把不同框架的资产接进平台覆盖率）。

两条主线：
1. import_official_cases(xlsx)  —— 官方测试报告 Excel → 手工用例库（覆盖率分母）。
   一个 app 一个 sheet（酷我/喜马拉雅/乐听/爱奇艺），逐行落 cases，用例编号 = 前缀-4位序号。
2. import_media_bundle(zip)     —— Media_automation 包的适配器：
   识别包内 docs 的映射 CSV（excel_id ↔ nodeid），把它的自动化用例接进覆盖率分子。

设计要点：平台不强推自己的标记约定，而是每种框架写一个「适配器」把它的映射喂进
统一的 script_case 表 —— 覆盖率计算、看板展示对框架来源无感知。
"""
import csv
import io
import re
import zipfile
from contextlib import closing
from datetime import datetime

from openpyxl import load_workbook

from .. import config, db

# 官方 Excel 每个 app sheet 的表头列序（0基）：编号/需求ID/模块/用例标题/前置/优先级
_COL_NO, _COL_REQ, _COL_MODULE, _COL_TITLE, _COL_PRE, _COL_PRIO = 0, 1, 2, 3, 4, 5

# Media_automation 的映射 CSV 识别特征：文件名含 mapping 且列含这些字段
_MAPPING_COLS = {"excel_id", "automation_case", "status"}

# 视为「已自动化」的状态（映射 CSV 的 status 字段）
_AUTOMATED_STATUS = {"AUTOMATED", "PARTIAL", "XFAIL_KNOWN_DEFECT"}


def case_id_for(prefix: str, excel_no) -> str:
    """统一的用例编号规则：前缀-4位序号（KW-0001）。各框架映射都按此拼编号。"""
    return f"{prefix}-{int(str(excel_no).strip()):04d}"


# ---------------------------------------------------------------- 官方分母导入
def import_official_cases(xlsx_bytes: bytes, replace: bool = True) -> dict:
    """官方 Excel → 手工用例库。

    replace=True：整库替换（清空 cases + 所有映射），把种子分母换成官方真实分母；
    这是「用官方 Excel 替换种子数据」的直接实现。返回各 app 导入条数。
    """
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception:
        raise ValueError("Excel 解析失败，请确认是官方测试报告 .xlsx")

    parsed = {}   # app全名 -> [(case_id, module, priority, title, req_id), ...]
    for sheet, (app, prefix) in config.OFFICIAL_APP_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        rows = []
        for i, row in enumerate(wb[sheet].iter_rows(values_only=True)):
            if i == 0:                          # 跳过表头
                continue
            cells = list(row) + [None] * 6
            no, req, module, title, _pre, prio = cells[:6]
            if no is None or title is None:
                continue
            try:
                int(str(no).strip())            # 编号必须是整数序号
            except (ValueError, TypeError):
                continue
            priority = str(prio).strip() if prio else "P3"
            if not re.match(r"^P\d$", priority):   # 优先级必须是 P1..P4（排除结构异常行）
                continue
            rows.append((case_id_for(prefix, no), str(module).strip() if module else "",
                         priority, str(title).strip(),
                         str(req).strip() if req else ""))
        if rows:
            parsed[app] = rows

    if not parsed:
        raise ValueError("未在 Excel 中识别到任何 app sheet（酷我/喜马拉雅/乐听/爱奇艺）")

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with closing(db.get_conn()) as conn:
        if replace:
            # 整库替换：官方成为唯一分母来源。种子脚本/映射一并清掉，避免悬空引用。
            conn.execute("DELETE FROM script_case")
            conn.execute("DELETE FROM scripts")
            conn.execute("DELETE FROM cases")
        counts = {}
        for app, rows in parsed.items():
            for cid, module, priority, title, req in rows:
                conn.execute(
                    "INSERT INTO cases(id,app,module,priority,title,req_id,source,created_at)"
                    " VALUES(?,?,?,?,?,?, 'official', ?)"
                    " ON CONFLICT(id) DO UPDATE SET app=excluded.app, module=excluded.module,"
                    " priority=excluded.priority, title=excluded.title, req_id=excluded.req_id,"
                    " source='official'",
                    (cid, app, module, priority, title, req, ts))
            counts[app] = len(rows)
        conn.commit()
    return {"apps": counts, "total": sum(counts.values()), "replaced": replace}


# ---------------------------------------------------------------- 框架适配器
def _find_mapping_csv(zf: zipfile.ZipFile):
    """在包内找 Media_automation 的映射 CSV（按文件名 + 列特征识别）。"""
    for info in zf.infolist():
        n = info.filename.lower()
        if n.endswith(".csv") and "mapping" in n:
            try:
                head = zf.read(info).decode("utf-8-sig", errors="replace").splitlines()[:1]
                if head and _MAPPING_COLS.issubset({c.strip() for c in head[0].split(",")}):
                    return info.filename
            except Exception:
                continue
    return None


def is_media_bundle(zip_bytes: bytes) -> bool:
    """判断上传的 zip 是否为 Media_automation 包（含可识别的映射 CSV）。"""
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            return _find_mapping_csv(zf) is not None
    except zipfile.BadZipFile:
        return False


def _parse_media_mapping(zip_bytes: bytes) -> tuple:
    """解析 Media 包映射 CSV，返回 (命中官方库的映射快照, 已自动化行数)。

    映射规则：CSV 每行 excel_id + status，凡 status 属于已自动化且对应官方用例
    （KW-{excel_id}）存在于 cases 表，即计入。只映射真实存在的用例，防虚高。
    """
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        csv_name = _find_mapping_csv(zf)
        if not csv_name:
            raise ValueError("未在包内找到映射 CSV，无法按 Media_automation 适配")
        text = zf.read(csv_name).decode("utf-8-sig", errors="replace")

    wanted, automated_rows = [], 0
    for r in csv.DictReader(io.StringIO(text)):
        if (r.get("status") or "").strip() not in _AUTOMATED_STATUS:
            continue
        if not (r.get("automation_case") or "").strip():
            continue
        automated_rows += 1
        wanted.append([case_id_for("KW", r["excel_id"]),
                       (r.get("module") or "").strip(), (r.get("priority") or "").strip()])
    with closing(db.get_conn()) as conn:
        existing = {r["id"] for r in conn.execute("SELECT id FROM cases WHERE app = ?", ("酷我音乐",))}
    return [w for w in wanted if w[0] in existing], automated_rows


def import_media_bundle(zip_bytes: bytes, name: str, version: str,
                        orig_filename: str = "bundle.zip") -> dict:
    """Media_automation 适配器：解析映射 CSV → 走多版本统一入口落库（save_version）。"""
    from . import script_service
    mapped, automated_rows = _parse_media_mapping(zip_bytes)
    r = script_service.save_version(name, "酷我音乐", version, config.FRAMEWORK_MEDIA,
                                    orig_filename, zip_bytes, mapped)
    return {**r, "mapped": len(mapped), "unmatched": automated_rows - len(mapped)}


# ---------------------------------------------------------------- Zcode 适配器
def _find_matrix_md(zf: zipfile.ZipFile):
    """在包内找 Zcode 的覆盖矩阵 markdown（文件名含覆盖矩阵/coverage + 内容像映射表）。"""
    for info in zf.infolist():
        n = info.filename
        if n.endswith(".md") and ("覆盖矩阵" in n or "coverage" in n.lower()):
            txt = zf.read(info).decode("utf-8", errors="replace")
            if "Excel 编号" in txt or "自动化用例" in txt:
                return txt
    return None


def is_zcode_bundle(zip_bytes: bytes) -> bool:
    """判断是否为 Media_automation_Zcode 包（覆盖矩阵 markdown + config/config.yaml）。"""
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            has_yaml = any(n.replace("\\", "/").endswith("config/config.yaml")
                           for n in zf.namelist())
            return has_yaml and _find_matrix_md(zf) is not None
    except zipfile.BadZipFile:
        return False


def _parse_zcode_mapping(zip_bytes: bytes) -> tuple:
    """解析 Zcode 覆盖矩阵 markdown，返回 (命中官方库的映射快照, 矩阵去重编号数)。

    markdown 表把「自动化用例 ↔ Excel 编号」列出来；一个 Excel 编号即官方用例
    KW-{编号}。只映射到官方库真实存在的用例。
    """
    from . import zcode_runner
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        md_text = _find_matrix_md(zf)
        if not md_text:
            raise ValueError("未在包内找到覆盖矩阵 markdown，无法按 Zcode 适配")
    entries = zcode_runner.parse_matrix(md_text)              # (file, func, param, excel_id)
    wanted_ids = {case_id_for("KW", e[3]) for e in entries}   # 去重的官方用例编号
    with closing(db.get_conn()) as conn:
        meta = {r["id"]: (r["module"], r["priority"]) for r in
                conn.execute("SELECT id, module, priority FROM cases WHERE app = ?", ("酷我音乐",))}
    return [[cid, *meta[cid]] for cid in sorted(wanted_ids) if cid in meta], len(wanted_ids)


def import_zcode_bundle(zip_bytes: bytes, name: str, version: str,
                        orig_filename: str = "bundle.zip") -> dict:
    """Zcode 适配器：解析覆盖矩阵 markdown → 走多版本统一入口落库（save_version）。"""
    from . import script_service
    mapped, distinct = _parse_zcode_mapping(zip_bytes)
    r = script_service.save_version(name, "酷我音乐", version, config.FRAMEWORK_ZCODE,
                                    orig_filename, zip_bytes, mapped)
    return {**r, "mapped": len(mapped), "unmatched": distinct - len(mapped)}


def inspect_zip(content: bytes) -> dict:
    """上传前只读识别（弹窗选完文件的摘要展示）：框架类型 + 映射命中统计，不落盘。"""
    if is_zcode_bundle(content):
        mapped, distinct = _parse_zcode_mapping(content)
        return {"framework": config.FRAMEWORK_ZCODE, "matched": len(mapped),
                "unmatched": distinct - len(mapped)}
    if is_media_bundle(content):
        mapped, automated = _parse_media_mapping(content)
        return {"framework": config.FRAMEWORK_MEDIA, "matched": len(mapped),
                "unmatched": automated - len(mapped)}
    return {}
